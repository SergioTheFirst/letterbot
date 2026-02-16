import io

import pytest

pytest.importorskip("openpyxl")

from openpyxl import Workbook

from mailbot_v26.bot_core.extractors.excel import extract_excel_text


def _build_workbook_bytes(fill_rows: int, fill_cols: int) -> bytes:
    buffer = io.BytesIO()
    workbook = Workbook()
    sheet = workbook.active
    for row in range(1, fill_rows + 1):
        for col in range(1, fill_cols + 1):
            sheet.cell(row=row, column=col, value=f"R{row}C{col}")
    workbook.save(buffer)
    return buffer.getvalue()


def test_extract_excel_text_openpyxl_values():
    buffer = io.BytesIO()
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Alpha"
    sheet["B1"] = "Beta"
    workbook.save(buffer)

    text = extract_excel_text(buffer.getvalue(), "sample.xlsx")

    assert "Alpha" in text
    assert "Beta" in text


def test_extract_excel_text_row_limit():
    data = _build_workbook_bytes(fill_rows=300, fill_cols=1)

    text = extract_excel_text(data, "rows.xlsx")
    lines = [line for line in text.splitlines() if line.strip()]

    assert len(lines) <= 200
    assert "R1C1" in text
    assert "R200C1" in text
    assert "R201C1" not in text


def test_extract_excel_text_column_limit():
    data = _build_workbook_bytes(fill_rows=1, fill_cols=35)

    text = extract_excel_text(data, "cols.xlsx")
    first_line = text.splitlines()[0]
    columns = first_line.split("\t")

    assert len(columns) <= 30
    assert "R1C30" in first_line
    assert "R1C31" not in first_line
