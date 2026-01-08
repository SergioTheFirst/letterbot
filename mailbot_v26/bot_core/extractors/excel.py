from __future__ import annotations

import io
import logging

logger = logging.getLogger(__name__)
MAX_ROWS = 200
MAX_COLS = 30
MAX_CHARS = 50_000


def _normalize_text(value: str) -> str:
    normalized = (value or "").encode("utf-8", "ignore").decode("utf-8", "ignore")
    return normalized.strip()


def _safe_text(value: str | None) -> str:
    if not value:
        return ""
    return " ".join(value.split())


def _stringify_cell(value: object | None) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return _safe_text(value)
    return _safe_text(str(value))


def _append_line(
    lines: list[str],
    line: str,
    *,
    max_chars: int = MAX_CHARS,
) -> bool:
    if not line:
        return True
    current = sum(len(item) for item in lines)
    if lines:
        current += len(lines) - 1
    remaining = max_chars - current
    if remaining <= 0:
        return False
    if len(line) > remaining:
        lines.append(line[:remaining])
        return False
    lines.append(line)
    return True


def _collect_rows_from_openpyxl(file_bytes: bytes) -> list[str]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        logger.debug("openpyxl import failed: %s", exc)
        return []

    try:
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        logger.debug("openpyxl load failed: %s", exc)
        return []

    lines: list[str] = []
    row_count = 0
    for sheet in workbook.worksheets:
        if row_count >= MAX_ROWS:
            break
        try:
            for row in sheet.iter_rows(max_col=MAX_COLS, values_only=True):
                if row_count >= MAX_ROWS:
                    break
                values = [_stringify_cell(cell) for cell in row[:MAX_COLS]]
                while values and not values[-1]:
                    values.pop()
                if not any(values):
                    continue
                line = "\t".join(values)
                if not _append_line(lines, line):
                    workbook.close()
                    return lines
                row_count += 1
        except Exception as exc:
            logger.debug("openpyxl sheet parse failed: %s", exc)
            continue
    workbook.close()
    return lines


def _collect_rows_from_xlrd(file_bytes: bytes) -> list[str]:
    try:
        import xlrd
    except Exception as exc:
        logger.debug("xlrd import failed: %s", exc)
        return []

    try:
        workbook = xlrd.open_workbook(file_contents=file_bytes)
    except Exception as exc:
        logger.debug("xlrd load failed: %s", exc)
        return []

    lines: list[str] = []
    row_count = 0
    for sheet in workbook.sheets():
        if row_count >= MAX_ROWS:
            break
        for rowx in range(min(sheet.nrows, MAX_ROWS - row_count)):
            if row_count >= MAX_ROWS:
                break
            values = []
            max_col = min(sheet.ncols, MAX_COLS)
            for colx in range(max_col):
                values.append(_stringify_cell(sheet.cell_value(rowx, colx)))
            while values and not values[-1]:
                values.pop()
            if not any(values):
                continue
            line = "\t".join(values)
            if not _append_line(lines, line):
                return lines
            row_count += 1
    return lines


def extract_excel(file_bytes: bytes, filename: str) -> str:
    """
    Извлечение из XLS/XLSX.
    Приоритет: openpyxl -> xlrd (если доступен)
    """
    name = (filename or "").lower()

    if not name.endswith((".xls", ".xlsx")):
        return ""

    # ПОПЫТКА 1: openpyxl (XLSX)
    if name.endswith(".xlsx"):
        rows = _collect_rows_from_openpyxl(file_bytes)
        if rows:
            text = _normalize_text("\n".join(rows))
            logger.debug("Extracted %d chars from %s (openpyxl)", len(text), filename)
            return text

    # ПОПЫТКА 2: xlrd (XLS, опционально)
    if name.endswith(".xls"):
        rows = _collect_rows_from_xlrd(file_bytes)
        if rows:
            text = _normalize_text("\n".join(rows))
            logger.debug("Extracted %d chars from %s (xlrd)", len(text), filename)
            return text

    logger.warning("Failed to extract text from %s", filename)
    return ""


extract_excel_text = extract_excel
